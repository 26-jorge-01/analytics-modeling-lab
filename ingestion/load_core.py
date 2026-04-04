import os
import logging
import pandas as pd
import csv
import io
from sqlalchemy import create_engine, text
from pathlib import Path
from typing import Optional, List, Dict, Any, Union, Iterator

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# Database Configuration (Defaults for local Docker environment)
DB_USER = os.getenv("POSTGRES_USER", "demo")
DB_PASSWORD = os.getenv("POSTGRES_PASSWORD", "demo")
DB_HOST = os.getenv("POSTGRES_HOST", "localhost")
DB_PORT = os.getenv("POSTGRES_PORT", "15432")
DB_NAME = os.getenv("POSTGRES_DB", "modeling_lab")

RAW_SCHEMA = "raw"

def get_engine():
    """Creates a SQLAlchemy engine for the Postgres database."""
    conn_str = (
        f"postgresql://{DB_USER}:{DB_PASSWORD}@"
        f"{DB_HOST}:{DB_PORT}/{DB_NAME}"
    )
    return create_engine(conn_str)

def ensure_schema_exists(engine, schema_name: str = RAW_SCHEMA):
    """Ensures that the specified schema exists in the database."""
    with engine.connect() as conn:
        conn.execute(text(f"CREATE SCHEMA IF NOT EXISTS {schema_name}"))
        conn.commit()
    logger.debug(f"Schema '{schema_name}' verified/created.")

def clean_table_name(filename: str) -> str:
    """
    Cleans the filename to derive a clean table name.
    """
    name = Path(filename).stem
    clean_name = name.lower().replace(" ", "_").replace("-", "_").replace(".", "_")
    # Specific project cleanup if needed
    clean_name = clean_name.replace("olist_", "").replace("_dataset", "")
    return clean_name

def psql_insert_copy(table, conn, keys, data_iter):
    """
    Execute SQL statement inserting data using COPY command.
    Bypasses standard INSERT for massive speed.
    """
    try:
        # Compatibility handling for SQLAlchemy 2.0
        if hasattr(conn, 'connection') and hasattr(conn.connection, 'dbapi_connection'):
            # SQLAlchemy 2.0
            dbapi_conn = conn.connection.dbapi_connection
        elif hasattr(conn, 'connection'):
            # SQLAlchemy 1.4
            dbapi_conn = conn.connection
        else:
            # Fallback
            dbapi_conn = conn.engine.raw_connection()

        with dbapi_conn.cursor() as cur:
            s_buf = io.StringIO()
            writer = csv.writer(s_buf)
            
            # Clean data for COPY command
            for row in data_iter:
                cleaned_row = []
                for val in row:
                    # Fix: If value is 15.0 and target is BIGINT, Postgres fails.
                    # Convert integer-like floats to actual ints for the CSV buffer.
                    if isinstance(val, float) and val.is_integer():
                        cleaned_row.append(int(val))
                    elif pd.isna(val):
                        # Use empty string for NULL in CSV mode
                        cleaned_row.append("")
                    else:
                        cleaned_row.append(val)
                writer.writerow(cleaned_row)
                
            s_buf.seek(0)

            columns = ', '.join(['"{}"'.format(k) for k in keys])
            if table.schema:
                table_name = '"{}"."{}"'.format(table.schema, table.name)
            else:
                table_name = '"{}"'.format(table.name)

            sql = 'COPY {} ({}) FROM STDIN WITH CSV'.format(
                table_name, columns)
            cur.copy_expert(sql=sql, file=s_buf)
            
    except Exception as e:
        logger.error(f"Error in psql_insert_copy: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise e

def dataframe_to_postgres(
    df: pd.DataFrame, 
    table_name: str, 
    schema: str = RAW_SCHEMA, 
    if_exists: str = "replace",
    batch_index: int = 0
):
    """
    Loads a dataframe to Postgres. 
    Handles 'replace' carefully by using TRUNCATE on the first batch.
    Uses high-speed COPY method.
    """
    engine = get_engine()
    ensure_schema_exists(engine, schema)
    
    # COMPATIBILITY FIX: Postgres truncates identifiers to 63 chars.
    # We must truncate our DataFrame columns to match Postgres logic, 
    # otherwise we get false "schema mismatches".
    df.columns = [str(c)[:63] for c in df.columns]
    
    actual_if_exists = if_exists
    if if_exists == "replace":
        if batch_index == 0:
            # Check if table exists to compare schema or truncate
            from sqlalchemy import inspect
            inspector = inspect(engine)
            if inspector.has_table(table_name, schema=schema):
                # Get existing columns (which are already truncated by Postgres)
                existing_columns = [col['name'] for col in inspector.get_columns(table_name, schema=schema)]
                df_columns = df.columns.tolist()
                
                # If columns match exactly, we can TRUNCATE (safe for dependencies)
                if set(existing_columns) == set(df_columns):
                    logger.info(f"Table {schema}.{table_name} matches schema. Truncating.")
                    with engine.connect() as conn:
                        conn.execute(text(f'TRUNCATE TABLE "{schema}"."{table_name}"'))
                        conn.commit()
                    actual_if_exists = "append"
                else:
                    # If columns DON'T match, we MUST drop and recreate
                    logger.warning(f"Schema mismatch for {table_name}. Falling back to full REPLACE with CASCADE.")
                    with engine.connect() as conn:
                        # Use CASCADE to handle dependent views
                        conn.execute(text(f'DROP TABLE IF EXISTS "{schema}"."{table_name}" CASCADE'))
                        conn.commit()
                    actual_if_exists = "replace"
            else:
                actual_if_exists = "replace"
        else:
            # For subsequent batches in a 'replace' operation, always append
            actual_if_exists = "append"

    df.to_sql(
        name=table_name,
        con=engine,
        schema=schema,
        if_exists=actual_if_exists,
        index=False,
        chunksize=50000,
        method=psql_insert_copy if engine.dialect.name == 'postgresql' else None
    )
    
    if batch_index == 0:
        logger.info(f"Started loading into {schema}.{table_name} using COPY protocol.")
    
def _stream_excel_chunks(file_path: Path, chunk_size: int, **kwargs) -> Iterator[pd.DataFrame]:
    """
    Streams an Excel file in chunks to save memory.
    Uses openpyxl in read_only mode.
    """
    import openpyxl
    
    # Extract sheet_name from kwargs or default to first
    sheet_name = kwargs.get('sheet_name', 0)
    skiprows = kwargs.get('skiprows', 0)
    
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb[sheet_name]
        
    # Use iter_rows to get all content
    rows_iter = ws.iter_rows(values_only=True)
    
    # Skip rows
    for _ in range(skiprows):
        next(rows_iter, None)
        
    # Get headers and find the real end of the data (strip trailing None columns)
    header_row = next(rows_iter, None)
    if not header_row:
        return
        
    # Find last non-None index in header
    last_col_idx = 0
    for i, val in enumerate(header_row):
        if val is not None:
            last_col_idx = i
            
    # Clean headers up to last_col_idx
    headers = []
    for i in range(last_col_idx + 1):
        val = header_row[i]
        # Match pandas 'Unnamed: N' naming convention and truncate to 63 chars for Postgres
        header_name = str(val) if val is not None else f"Unnamed: {i}"
        headers.append(header_name[:63])
    
    logger.info(f"Detected {len(headers)} columns in Excel file (after stripping empty trailing columns).")

    chunk = []
    for row in rows_iter:
        # Truncate row to match headers length
        truncated_row = row[:last_col_idx + 1]
        chunk.append(truncated_row)
        
        if len(chunk) >= chunk_size:
            yield pd.DataFrame(chunk, columns=headers)
            chunk = []
            
    if chunk:
        yield pd.DataFrame(chunk, columns=headers)
        
    wb.close()

def load_file(
    file_path: Union[str, Path], 
    table_name: Optional[str] = None, 
    schema: str = RAW_SCHEMA,
    if_exists: str = "replace",
    chunk_size: Optional[int] = None,
    **kwargs
) -> bool:
    """
    General function to load a file (CSV, Excel, Parquet) to Postgres.
    Support chunked loading for CSV and Excel.
    """
    path = Path(file_path)
    if not path.exists():
        logger.error(f"File not found: {path}")
        return False

    ext = path.suffix.lower()
    table_name = table_name or clean_table_name(path.name)
    
    try:
        logger.info(f"Loading {path.name} into {schema}.{table_name} (Chunk size: {chunk_size})...")
        
        if ext == '.csv':
            if chunk_size:
                reader = pd.read_csv(path, chunksize=chunk_size, **kwargs)
                for i, df in enumerate(reader):
                    dataframe_to_postgres(df, table_name, schema, if_exists, batch_index=i)
                logger.info(f"Finished loading {path.name} in chunks.")
            else:
                df = pd.read_csv(path, **kwargs)
                dataframe_to_postgres(df, table_name, schema, if_exists)
                
        elif ext in ['.xlsx', '.xls']:
            if chunk_size and ext == '.xlsx':
                # Use streaming for .xlsx
                for i, df in enumerate(_stream_excel_chunks(path, chunk_size, **kwargs)):
                    dataframe_to_postgres(df, table_name, schema, if_exists, batch_index=i)
                logger.info(f"Finished loading {path.name} using streaming chunks.")
            else:
                # Standard pandas read for small files or .xls
                df = pd.read_excel(path, **kwargs)
                dataframe_to_postgres(df, table_name, schema, if_exists)
                
        elif ext == '.parquet':
            # Parquet is usually efficient, but we can still chunk if needed 
            # (though standard pandas doesn't support chunksize for read_parquet as easily)
            df = pd.read_parquet(path, **kwargs)
            dataframe_to_postgres(df, table_name, schema, if_exists)
        else:
            logger.error(f"Unsupported file format: {ext}")
            return False

        return True

    except Exception as e:
        logger.error(f"FATAL: Error loading {path.name}: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def load_directory(
    directory_path: Union[str, Path], 
    extension: str = "*.xlsx", 
    schema: str = RAW_SCHEMA,
    if_exists: str = "replace",
    chunk_size: Optional[int] = None,
    **kwargs
) -> Dict[str, bool]:
    """
    Loads all files with a specific extension from a directory.
    """
    path = Path(directory_path)
    if not path.is_dir():
        logger.error(f"Directory not found: {path}")
        return {}

    files = list(path.glob(extension))
    logger.info(f"Found {len(files)} files with extension {extension} in {path}")
    
    results = {}
    for f in files:
        success = load_file(f, schema=schema, if_exists=if_exists, chunk_size=chunk_size, **kwargs)
        results[f.name] = success
        
    return results